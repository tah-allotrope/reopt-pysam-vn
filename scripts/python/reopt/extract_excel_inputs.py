"""
Extract Saigon18 project data from Excel model for use as REopt inputs.

Usage:
    python scripts/python/reopt/extract_excel_inputs.py \
        --excel "path/to/llm 20260129 SOLAR BESS MODEL - Editing - for processing test.xlsx" \
        --output data/interim/saigon18/2026-03-20_saigon18_extracted_inputs.json
"""

import argparse
import json
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PV_KW_RATED = 40_360.0  # fixed design capacity (kW)
DATA_START_ROW = 9  # first data row in "Data Input" sheet (1-indexed)
DATA_END_ROW = 8768  # last data row (row 9 + 8759 = row 8768)


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------


def _safe_float(value, default: float = 0.0) -> float:
    """Return float(value) or default when value is None or non-numeric."""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def extract_data_input(data_ws) -> dict:
    """Extract 8760-hour hourly profiles from the 'Data Input' sheet.

    Expected column layout (0-indexed within each row tuple):
        col B (index 1): SimulationProfile_kW  — PV generation
        col D (index 3): Load_kW               — factory load
        col E (index 4): FMP (VND/MWh)
        col F (index 5): CFMP (VND/MWh)
    """
    rows = list(
        data_ws.iter_rows(
            min_row=DATA_START_ROW, max_row=DATA_END_ROW, values_only=True
        )
    )
    if len(rows) != 8760:
        raise ValueError(
            f"Expected 8760 data rows in 'Data Input' sheet, got {len(rows)}. "
            f"Check DATA_START_ROW ({DATA_START_ROW}) and sheet structure."
        )

    pv_kw_raw = [_safe_float(r[1]) for r in rows]  # col B
    loads_kw = [_safe_float(r[3]) for r in rows]  # col D
    fmp = [_safe_float(r[4]) for r in rows]  # col E
    cfmp = [_safe_float(r[5]) for r in rows]  # col F

    pv_prod_factor = [v / PV_KW_RATED for v in pv_kw_raw]

    return {
        "loads_kw": loads_kw,
        "pv_kw_raw": pv_kw_raw,
        "pv_production_factor_series": pv_prod_factor,
        "fmp_vnd_per_mwh": fmp,
        "cfmp_vnd_per_mwh": cfmp,
    }


def extract_assumption(assump_ws) -> dict:
    """Extract key scalar parameters from the 'Assumption' sheet.

    Reads a block of rows and attempts to locate known parameters.
    Falls back to plan-documented design values when cells are blank.
    All row references are 1-indexed as in Excel.
    """
    # Read a wide enough block; plan docs reference up to row 69.
    raw = {}
    for row in assump_ws.iter_rows(min_row=1, max_row=70, values_only=True):
        pass  # intentional — actual lookups done via cell() below

    def cell(row: int, col: int):
        """Read a single cell value (1-indexed)."""
        return assump_ws.cell(row=row, column=col).value

    # Map Excel row → parameter (col B = column 2, col P = column 16)
    solar_kw = _safe_float(cell(15, 2), PV_KW_RATED) * 1_000  # MWp → kW
    # Row 15 col B stores MWp; handle both MWp (<100) and kW (>1000)
    if solar_kw < 1_000:
        solar_kw = solar_kw * 1_000  # was in MWp, convert to kW
    if solar_kw == 0:
        solar_kw = PV_KW_RATED  # fallback

    pr = _safe_float(cell(16, 2), 0.8086)  # performance ratio
    annual_yield_gwh = _safe_float(cell(18, 2), 71.808)  # GWh — may be in MWh
    spec_energy = _safe_float(cell(19, 2), 1779.0)  # kWh/kWp

    bess_kwh = _safe_float(cell(25, 2), 66_000.0)
    bess_kw = _safe_float(cell(26, 2), 20_000.0)
    bess_dod = _safe_float(cell(27, 2), 0.85)
    half_cycle_eff = _safe_float(cell(28, 2), 0.95)
    bess_degradation = _safe_float(cell(29, 2), 0.03)

    # Financial (col O = column 15, col P = column 16)
    analysis_years = int(_safe_float(cell(10, 15), 20))
    om_pv_per_kw = _safe_float(cell(25, 15), 6.0)  # $/kW/yr
    om_bess_per_kwh = _safe_float(cell(27, 15), 2.0)  # $/kWh/yr
    opex_esc = _safe_float(cell(34, 15), 0.04)
    cit_rate = _safe_float(cell(62, 15), 0.20)

    # Tariff (col P = column 16)
    tariff_standard_vnd = _safe_float(cell(14, 16), 1_811.0)
    tariff_peak_vnd = _safe_float(cell(15, 16), 3_266.0)
    tariff_offpeak_vnd = _safe_float(cell(16, 16), 1_146.0)
    ppa_discount = _safe_float(cell(30, 15), 0.15)  # Assumption!O30

    # CAPEX (col B rows 41-42)
    pv_capex_per_kw = _safe_float(cell(41, 2), 750.0)
    bess_capex_per_kwh = _safe_float(cell(42, 2), 200.0)

    return {
        "solar_kw_rated": solar_kw if solar_kw >= 1_000 else PV_KW_RATED,
        "performance_ratio": pr,
        "annual_yield_gwh": annual_yield_gwh,
        "specific_energy_kwh_per_kwp": spec_energy,
        "bess_kwh": bess_kwh,
        "bess_kw": bess_kw,
        "bess_dod_fraction": bess_dod,
        "bess_half_cycle_efficiency": half_cycle_eff,
        "bess_degradation_per_year": bess_degradation,
        "analysis_years": analysis_years if 1 <= analysis_years <= 40 else 20,
        "om_pv_per_kw_per_year": om_pv_per_kw,
        "om_bess_per_kwh_per_year": om_bess_per_kwh,
        "opex_escalation": opex_esc,
        "cit_rate": cit_rate,
        "tariff_standard_vnd_per_kwh": tariff_standard_vnd,
        "tariff_peak_vnd_per_kwh": tariff_peak_vnd,
        "tariff_offpeak_vnd_per_kwh": tariff_offpeak_vnd,
        "ppa_discount_fraction": ppa_discount,
        "pv_capex_usd_per_kw": pv_capex_per_kw,
        "bess_capex_usd_per_kwh": bess_capex_per_kwh,
    }


def validate_extracted(profiles: dict, assumptions: dict) -> list[str]:
    """Return a list of validation warnings (not errors). Empty list = all OK."""
    warnings_list = []

    pv_raw = profiles["pv_kw_raw"]
    loads = profiles["loads_kw"]
    pv_prod = profiles["pv_production_factor_series"]

    # 8760 length checks
    for name, arr in [
        ("loads_kw", loads),
        ("pv_kw_raw", pv_raw),
        ("fmp", profiles["fmp_vnd_per_mwh"]),
    ]:
        if len(arr) != 8760:
            warnings_list.append(f"[FAIL] {name}: expected 8760 values, got {len(arr)}")

    # No negative loads
    neg_loads = sum(1 for v in loads if v < 0)
    if neg_loads:
        warnings_list.append(f"[WARN] loads_kw has {neg_loads} negative values")

    # No negative PV
    neg_pv = sum(1 for v in pv_raw if v < 0)
    if neg_pv:
        warnings_list.append(f"[WARN] pv_kw_raw has {neg_pv} negative values")

    # Production factors in [0, 1]
    bad_pf = sum(1 for v in pv_prod if v < 0 or v > 1.05)
    if bad_pf:
        warnings_list.append(
            f"[WARN] {bad_pf} production factors outside [0, 1.05] — "
            "check pv_kw_rated or dc_ac_ratio clipping"
        )

    # Annual solar yield vs. plan target (71.8 GWh ± 2%)
    annual_pv_gwh = sum(pv_raw) / 1e6
    target_gwh = 71.808
    if annual_pv_gwh > 0:
        delta_pct = abs(annual_pv_gwh - target_gwh) / target_gwh * 100
        if delta_pct > 2.0:
            warnings_list.append(
                f"[WARN] Annual PV generation {annual_pv_gwh:.2f} GWh differs from "
                f"plan target {target_gwh} GWh by {delta_pct:.1f}% (tolerance 2%)"
            )
    else:
        warnings_list.append(
            "[WARN] PV profile is all zeros — check column B in Data Input sheet"
        )

    # Annual load sanity check (plan: 184.3 GWh)
    annual_load_gwh = sum(loads) / 1e6
    if annual_load_gwh < 50 or annual_load_gwh > 300:
        warnings_list.append(
            f"[WARN] Annual load {annual_load_gwh:.1f} GWh is outside expected range [50, 300] GWh"
        )

    return warnings_list


def extract_saigon18_data(excel_path: str) -> dict:
    """Load the Excel workbook and return a consolidated dict of all extracted data."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_names = wb.sheetnames
    required_sheets = ["Data Input", "Assumption"]
    for s in required_sheets:
        if s not in sheet_names:
            raise ValueError(
                f"Required sheet '{s}' not found. Available sheets: {sheet_names}"
            )

    profiles = extract_data_input(wb["Data Input"])
    assumptions = extract_assumption(wb["Assumption"])

    warnings_list = validate_extracted(profiles, assumptions)

    pv_raw = profiles["pv_kw_raw"]
    loads = profiles["loads_kw"]

    result = {
        # Hourly profiles (8760 values each)
        "loads_kw": profiles["loads_kw"],
        "pv_production_factor_series": profiles["pv_production_factor_series"],
        "fmp_vnd_per_mwh": profiles["fmp_vnd_per_mwh"],
        "cfmp_vnd_per_mwh": profiles["cfmp_vnd_per_mwh"],
        # Scalar summaries
        "solar_kw_rated": assumptions.get("solar_kw_rated", PV_KW_RATED),
        "data_year": 2024,
        "location": "Vietnam (south, HCMC area)",
        "annual_solar_gwh": sum(pv_raw) / 1e6,
        "annual_load_gwh": sum(loads) / 1e6,
        "peak_load_kw": max(loads) if loads else 0,
        # Assumptions from Excel
        "assumptions": assumptions,
        # Validation
        "validation_warnings": warnings_list,
        "validation_passed": all("[FAIL]" not in w for w in warnings_list),
    }
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Extract Saigon18 Excel data to JSON")
    parser.add_argument("--excel", required=True, help="Path to Excel workbook")
    parser.add_argument(
        "--output",
        default="data/interim/saigon18/2026-03-20_saigon18_extracted_inputs.json",
        help="Output JSON path (default: data/interim/saigon18/2026-03-20_saigon18_extracted_inputs.json)",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    print(f"Loading: {excel_path}")
    data = extract_saigon18_data(str(excel_path))

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print(f"Saved to: {output_path}")
    print(f"  Annual PV generation : {data['annual_solar_gwh']:.2f} GWh")
    print(f"  Annual load          : {data['annual_load_gwh']:.2f} GWh")
    print(f"  Peak load            : {data['peak_load_kw']:.1f} kW")

    if data["validation_warnings"]:
        print("\nValidation warnings:")
        for w in data["validation_warnings"]:
            print(f"  {w}")
    else:
        print("\nValidation: all checks passed.")

    if not data["validation_passed"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
