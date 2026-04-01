"""
Rank case studies for pure physical load matching against a 30 MWp Ninh Thuan
solar plant with a BESS power cap of 6 MW.

The ranking is intentionally commercial-agnostic: it screens only how well each
candidate load profile can physically absorb the reference solar output,
optionally helped by a simple 6 MW instantaneous battery charging headroom
proxy during solar hours.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent.parent
CASE_STUDIES_DIR = REPO_ROOT / "scenarios" / "case_studies"
DEFAULT_JSON_OUT = (
    REPO_ROOT
    / "artifacts"
    / "reports"
    / "case_studies"
    / "2026-04-01_offtaker-physical-match-ranking.json"
)
DEFAULT_HTML_OUT = (
    REPO_ROOT / "reports" / "2026-04-01-case-study-offtaker-physical-match-ranking.html"
)
DEFAULT_TEMPLATE = (
    Path.home()
    / ".config"
    / "opencode"
    / "skills"
    / "report"
    / "assets"
    / "report-template.html"
)
REFERENCE_SOLAR_SCENARIO = (
    CASE_STUDIES_DIR / "saigon18" / "2026-03-20_scenario-a_fixed-sizing_evntou.json"
)

SOLAR_CAPACITY_KW = 30_000.0
BESS_POWER_KW = 6_000.0
REPORT_DATE = "2026-04-01"
PROJECT_NAME = "Vietnam Case Study Offtaker Screening"
REPO_NAME = REPO_ROOT.name

CASE_DEFINITIONS = [
    {
        "case": "north_thuan",
        "label": "North Thuan",
        "kind": "json",
        "path": CASE_STUDIES_DIR / "north_thuan" / "north_thuan_scenario_a.json",
        "notes": "Synthetic industrial 8760 built for the North Thuan 30 MW solar validation path.",
    },
    {
        "case": "ninhsim",
        "label": "Ninh Sim",
        "kind": "csv",
        "path": CASE_STUDIES_DIR / "ninhsim" / "NinhsimSample.csv",
        "notes": "Raw sampled industrial load CSV used in the Saigon18/Ninh Sim workstream.",
    },
    {
        "case": "saigon18",
        "label": "Saigon18",
        "kind": "json",
        "path": CASE_STUDIES_DIR
        / "saigon18"
        / "2026-03-20_scenario-a_fixed-sizing_evntou.json",
        "notes": "Extracted real-project 8760 load profile embedded in the Saigon18 baseline scenario.",
    },
    {
        "case": "verdant",
        "label": "Verdant",
        "kind": "csv",
        "path": CASE_STUDIES_DIR / "verdant" / "Verdant.csv",
        "notes": "Small intermittent facility load with short daytime spikes.",
    },
    {
        "case": "regina",
        "label": "Regina",
        "kind": "xlsx",
        "path": CASE_STUDIES_DIR / "regina" / "Regina.xlsx",
        "notes": "Single-column workbook containing one header row plus 8760 hourly loads.",
    },
    {
        "case": "emivest",
        "label": "Emivest",
        "kind": "csv",
        "path": CASE_STUDIES_DIR / "emivest" / "Emivest.csv",
        "notes": "Highly peaky profile with very low sustained daytime demand.",
    },
]


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "phase-report"


def clean_numeric(raw_value: object) -> float | None:
    if raw_value is None:
        return None

    text = str(raw_value).replace("\ufeff", "").strip()
    text = text.strip('"').strip()
    text = text.replace(",", "")
    if text in {"", "-", "NA", "N/A", "None", "null"}:
        return None

    return float(text)


def interpolate_missing(values: list[float | None]) -> tuple[list[float], dict]:
    filled = list(values)
    interpolated_indices = []

    for index, value in enumerate(filled):
        if value is not None:
            continue

        left_index = index - 1
        while left_index >= 0 and filled[left_index] is None:
            left_index -= 1

        right_index = index + 1
        while right_index < len(filled) and filled[right_index] is None:
            right_index += 1

        left_value = filled[left_index] if left_index >= 0 else None
        right_value = filled[right_index] if right_index < len(filled) else None

        if left_value is None and right_value is None:
            raise ValueError("Load series contains only missing values")
        if left_value is None:
            assert right_value is not None
            filled[index] = float(right_value)
        elif right_value is None:
            assert left_value is not None
            filled[index] = float(left_value)
        else:
            filled[index] = float(left_value + right_value) / 2.0

        interpolated_indices.append(index)

    return [float(value) for value in filled if value is not None], {
        "missing_count": len(interpolated_indices),
        "interpolated_indices": interpolated_indices,
    }


def sanitize_load_series(values: list[float | None]) -> tuple[list[float], dict]:
    clipped_negative_count = 0
    precleaned = []

    for value in values:
        if value is not None and value < 0:
            precleaned.append(0.0)
            clipped_negative_count += 1
        else:
            precleaned.append(value)

    filled, issues = interpolate_missing(precleaned)
    cleaned = []

    for value in filled:
        cleaned.append(float(max(value, 0.0)))

    issues["clipped_negative_count"] = clipped_negative_count
    issues["final_count"] = len(cleaned)
    return cleaned, issues


def read_csv_loads(path: Path) -> tuple[list[float], dict]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    values = [clean_numeric(row[0] if row else None) for row in rows[1:]]
    return sanitize_load_series(values)


def read_xlsx_loads(path: Path) -> tuple[list[float], dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    values = []
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index == 0:
            continue
        values.append(clean_numeric(row[0]))
    workbook.close()
    return sanitize_load_series(values)


def read_json_loads(path: Path) -> tuple[list[float], dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    values = [clean_numeric(value) for value in payload["ElectricLoad"]["loads_kw"]]
    return sanitize_load_series(values)


def load_reference_solar_profile(
    reference_path: Path = REFERENCE_SOLAR_SCENARIO,
) -> list[float]:
    payload = json.loads(reference_path.read_text(encoding="utf-8"))
    production_factors = payload["PV"].get("production_factor_series")
    if not production_factors:
        raise ValueError(
            f"Reference solar scenario {reference_path} does not contain PV.production_factor_series"
        )

    if len(production_factors) != 8760:
        raise ValueError("Reference solar profile must contain exactly 8760 hours")

    return [float(value) * SOLAR_CAPACITY_KW for value in production_factors]


def compute_fit_score(metrics: dict) -> float:
    annual_with_bess = metrics["solar_absorption_with_bess_pct"] / 100.0
    annual_direct = metrics["solar_absorption_no_bess_pct"] / 100.0
    full_hours_with_bess = metrics["solar_hours_fully_absorbed_with_bess_pct"] / 100.0
    solar_hour_floor = min(metrics["min_solar_hour_load_mw"] / 30.0, 1.0)
    return round(
        100.0
        * (
            0.55 * annual_with_bess
            + 0.20 * annual_direct
            + 0.15 * full_hours_with_bess
            + 0.10 * solar_hour_floor
        ),
        1,
    )


def build_rationale(metrics: dict) -> str:
    absorption = metrics["solar_absorption_with_bess_pct"]
    min_solar_hour_load = metrics["min_solar_hour_load_mw"]
    curtailment = metrics["curtailment_with_bess_gwh"]

    if absorption >= 98.0:
        opening = "Absorbs essentially all of the reference solar profile"
    elif absorption >= 90.0:
        opening = "Absorbs most of the reference solar profile"
    elif absorption >= 50.0:
        opening = "Absorbs only a moderate share of the reference solar profile"
    else:
        opening = "Leaves most of the reference solar profile unmatched"

    if min_solar_hour_load >= 24.0:
        floor_note = "solar-hour demand stays high enough that curtailment risk is structurally low"
    elif min_solar_hour_load >= 10.0:
        floor_note = "solar-hour demand is respectable but still needs the 6 MW BESS proxy to close part of the gap"
    else:
        floor_note = "solar-hour demand collapses too often for a 30 MWp plant to sit comfortably behind the meter"

    return (
        f"{opening} ({absorption:.1f}% with the 6 MW BESS proxy); "
        f"minimum solar-hour load is {min_solar_hour_load:.2f} MW and residual curtailment is {curtailment:.2f} GWh/yr, so {floor_note}."
    )


def summarize_case(
    case_definition: dict,
    load_series_kw: list[float],
    solar_series_kw: list[float],
    issues: dict,
) -> dict:
    if len(load_series_kw) != 8760:
        raise ValueError(
            f"{case_definition['case']} load series must have 8760 hours, got {len(load_series_kw)}"
        )

    direct_match = [
        min(load_kw, solar_kw)
        for load_kw, solar_kw in zip(load_series_kw, solar_series_kw, strict=True)
    ]
    bess_match = [
        min(load_kw + BESS_POWER_KW, solar_kw)
        for load_kw, solar_kw in zip(load_series_kw, solar_series_kw, strict=True)
    ]
    curtailed_direct = [
        max(solar_kw - load_kw, 0.0)
        for load_kw, solar_kw in zip(load_series_kw, solar_series_kw, strict=True)
    ]
    curtailed_bess = [
        max(solar_kw - (load_kw + BESS_POWER_KW), 0.0)
        for load_kw, solar_kw in zip(load_series_kw, solar_series_kw, strict=True)
    ]
    solar_hour_indices = [
        index for index, solar_kw in enumerate(solar_series_kw) if solar_kw > 0.0
    ]

    solar_hour_loads = [load_series_kw[index] for index in solar_hour_indices]
    solar_hour_solar = [solar_series_kw[index] for index in solar_hour_indices]
    full_direct_hours = sum(
        1
        for load_kw, solar_kw in zip(solar_hour_loads, solar_hour_solar, strict=True)
        if load_kw >= solar_kw
    )
    full_bess_hours = sum(
        1
        for load_kw, solar_kw in zip(solar_hour_loads, solar_hour_solar, strict=True)
        if load_kw + BESS_POWER_KW >= solar_kw
    )

    annual_load_kwh = sum(load_series_kw)
    solar_annual_kwh = sum(solar_series_kw)
    direct_match_kwh = sum(direct_match)
    bess_match_kwh = sum(bess_match)

    metrics = {
        "case": case_definition["case"],
        "label": case_definition["label"],
        "source_kind": case_definition["kind"],
        "source_path": str(case_definition["path"].relative_to(REPO_ROOT)),
        "notes": case_definition["notes"],
        "annual_load_gwh": annual_load_kwh / 1_000_000.0,
        "average_load_mw": annual_load_kwh / 8760.0 / 1000.0,
        "peak_load_mw": max(load_series_kw) / 1000.0,
        "minimum_load_mw": min(load_series_kw) / 1000.0,
        "load_factor_pct": 100.0 * (sum(load_series_kw) / 8760.0) / max(load_series_kw),
        "solar_annual_gwh": solar_annual_kwh / 1_000_000.0,
        "direct_match_gwh": direct_match_kwh / 1_000_000.0,
        "matched_with_bess_gwh": bess_match_kwh / 1_000_000.0,
        "curtailment_no_bess_gwh": sum(curtailed_direct) / 1_000_000.0,
        "curtailment_with_bess_gwh": sum(curtailed_bess) / 1_000_000.0,
        "solar_absorption_no_bess_pct": 100.0 * direct_match_kwh / solar_annual_kwh,
        "solar_absorption_with_bess_pct": 100.0 * bess_match_kwh / solar_annual_kwh,
        "solar_hours_fully_absorbed_no_bess_pct": 100.0
        * full_direct_hours
        / len(solar_hour_indices),
        "solar_hours_fully_absorbed_with_bess_pct": 100.0
        * full_bess_hours
        / len(solar_hour_indices),
        "average_solar_hour_load_mw": sum(solar_hour_loads)
        / len(solar_hour_loads)
        / 1000.0,
        "min_solar_hour_load_mw": min(solar_hour_loads) / 1000.0,
        "cleaning": issues,
    }
    metrics["fit_score"] = compute_fit_score(metrics)
    metrics["rationale"] = build_rationale(metrics)
    return metrics


def load_case_metrics(case_definition: dict, solar_series_kw: list[float]) -> dict:
    loader_map = {
        "csv": read_csv_loads,
        "xlsx": read_xlsx_loads,
        "json": read_json_loads,
    }
    load_series_kw, issues = loader_map[case_definition["kind"]](
        case_definition["path"]
    )
    return summarize_case(case_definition, load_series_kw, solar_series_kw, issues)


def rank_cases(case_metrics: list[dict]) -> list[dict]:
    ranked = sorted(
        case_metrics,
        key=lambda row: (
            row["fit_score"],
            row["solar_absorption_with_bess_pct"],
            row["solar_absorption_no_bess_pct"],
        ),
        reverse=True,
    )
    for rank, row in enumerate(ranked, start=1):
        row["rank"] = rank
    return ranked


def build_results() -> dict:
    solar_series_kw = load_reference_solar_profile()
    ranked_cases = rank_cases(
        [
            load_case_metrics(case_definition, solar_series_kw)
            for case_definition in CASE_DEFINITIONS
        ]
    )

    return {
        "report_date": REPORT_DATE,
        "project": PROJECT_NAME,
        "repository": REPO_NAME,
        "screening_basis": {
            "mode": "pure_physical_load_matching",
            "solar_capacity_kw": SOLAR_CAPACITY_KW,
            "bess_power_cap_kw": BESS_POWER_KW,
            "reference_solar_profile": str(
                REFERENCE_SOLAR_SCENARIO.relative_to(REPO_ROOT)
            ),
            "battery_model": "Simple screening proxy: adds up to 6 MW of instantaneous solar-hour absorption headroom; no battery duration optimization is modeled.",
            "ranking_rule": "Rank by physical absorption of the reference solar profile, not by tariff, DPPA structure, or contract risk.",
        },
        "ranking": ranked_cases,
    }


def render_column(label: str, items: list[str]) -> str:
    bullets = "".join(f"<li>{item}</li>" for item in items)
    return (
        '<div class="column-card">'
        f'<h3 class="column-label">{label}</h3>'
        '<div class="splitter"></div>'
        f"<ul>{bullets}</ul>"
        "</div>"
    )


def render_tools_table(rows: list[tuple[str, str, str]]) -> str:
    body = "".join(
        f"<tr><td><code>{tool}</code></td><td>{purpose}</td><td>{outcome}</td></tr>"
        for tool, purpose, outcome in rows
    )
    return (
        "<table><thead><tr><th>Tool / Method</th><th>Purpose</th><th>Outcome</th></tr></thead>"
        f"<tbody>{body}</tbody></table>"
    )


def render_open_questions(items: list[str]) -> str:
    return "<ul>" + "".join(f"<li>{item}</li>" for item in items) + "</ul>"


def render_chart(
    canvas_id: str, chart_type: str, data: dict, options: dict | None = None
) -> str:
    chart_json = json.dumps(
        {"type": chart_type, "data": data, "options": options or {}},
        ensure_ascii=True,
    )
    return (
        f'<canvas id="{canvas_id}"></canvas>'
        "<script>"
        f"new Chart(document.getElementById('{canvas_id}'), {chart_json});"
        "</script>"
    )


def build_report_html(results: dict, template_text: str) -> str:
    ranked = results["ranking"]
    top_three = ranked[:3]
    ranking_items = [
        (
            f"#{row['rank']} {row['label']} - fit score {row['fit_score']:.1f}",
            f"Absorption with 6 MW BESS proxy {row['solar_absorption_with_bess_pct']:.1f}%, residual curtailment {row['curtailment_with_bess_gwh']:.2f} GWh/yr.",
        )
        for row in ranked
    ]
    ranking_bullets = [
        f"<strong>{title}</strong><br>{detail}" for title, detail in ranking_items
    ]
    cleaning_notes = []
    for row in ranked:
        issues = row["cleaning"]
        if issues["missing_count"] or issues["clipped_negative_count"]:
            cleaning_notes.append(
                f"{row['label']}: interpolated {issues['missing_count']} missing hour(s) and clipped {issues['clipped_negative_count']} negative hour(s) to zero."
            )

    if not cleaning_notes:
        cleaning_notes.append(
            "No source cleaning was required beyond normal numeric parsing."
        )

    input_output_html = render_column(
        "Input",
        [
            "Screened all six `scenarios/case_studies/` candidates against a common 30 MWp south-central Vietnam solar profile built from the embedded `PV.production_factor_series` in `scenarios/case_studies/saigon18/2026-03-20_scenario-a_fixed-sizing_evntou.json`.",
            "Used pure physical load matching only: no tariff, DPPA, strike-price, or contract-structure weighting entered the rank order.",
            "Applied a simple 6 MW BESS power-cap proxy as extra solar-hour absorption headroom, without trying to optimize battery duration or dispatch.",
        ],
    ) + render_column(
        "Output",
        [
            f"Ranked all six cases and wrote the machine-readable artifact to `{DEFAULT_JSON_OUT.relative_to(REPO_ROOT)}`.",
            f"Top three physical matches: {top_three[0]['label']}, {top_three[1]['label']}, and {top_three[2]['label']}.",
            f"Best-fit case `{ranked[0]['label']}` absorbs {ranked[0]['solar_absorption_with_bess_pct']:.1f}% of the reference solar profile with the 6 MW BESS proxy.",
        ],
    )

    tools_html = render_tools_table(
        [
            (
                "rank_case_study_offtakers.py",
                "Normalize mixed CSV/XLSX/JSON case-study inputs into one screening workflow",
                "Produced a reproducible physical-match ranking artifact and HTML phase report.",
            ),
            (
                "Reference solar profile replay",
                "Apply one south-central Vietnam 30 MWp hourly solar shape to every candidate load",
                "Made the comparison consistent across all six offtaker candidates.",
            ),
            (
                "6 MW BESS power-headroom proxy",
                "Approximate the limited battery's help without pretending to solve a full dispatch model",
                "Converted the user constraint into a transparent screening heuristic.",
            ),
            (
                "Source cleaning",
                "Repair sparse data quality issues before ranking",
                "Handled one `-` placeholder in `ninhsim` and one negative value in `emivest` without dropping hours.",
            ),
        ]
    )

    charts_html = render_chart(
        "caseStudyFitChart",
        "bar",
        {
            "labels": [row["label"] for row in ranked],
            "datasets": [
                {
                    "label": "Fit score",
                    "data": [row["fit_score"] for row in ranked],
                    "backgroundColor": [
                        "#39ff14",
                        "#00f5ff",
                        "#5fb9ff",
                        "#ffb100",
                        "#ff7a00",
                        "#ff4d6d",
                    ],
                },
                {
                    "label": "Solar absorption with 6 MW BESS proxy (%)",
                    "data": [row["solar_absorption_with_bess_pct"] for row in ranked],
                    "backgroundColor": "rgba(0, 245, 255, 0.35)",
                },
            ],
        },
        {"responsive": True, "plugins": {"legend": {"position": "bottom"}}},
    ) + render_chart(
        "caseStudyCurtailmentChart",
        "bar",
        {
            "labels": [row["label"] for row in ranked],
            "datasets": [
                {
                    "label": "Residual curtailment with 6 MW BESS proxy (GWh/yr)",
                    "data": [row["curtailment_with_bess_gwh"] for row in ranked],
                    "backgroundColor": [
                        "#39ff14",
                        "#00f5ff",
                        "#5fb9ff",
                        "#ffb100",
                        "#ff7a00",
                        "#ff4d6d",
                    ],
                }
            ],
        },
        {"responsive": True, "plugins": {"legend": {"position": "bottom"}}},
    )

    open_questions_html = render_open_questions(
        [
            "Battery duration is still abstracted away; the current screen treats the 6 MW cap as instantaneous charging headroom only, so a later phase could test explicit 2-hour or 4-hour BESS assumptions.",
            "The ranking is intentionally commercial-agnostic; if a later phase needs developer prioritization, add tariff, interconnection, and contract-structure filters separately rather than mixing them into this physical-fit rank.",
            *cleaning_notes,
        ]
    )

    replacements = {
        "{{PHASE_NAME}}": "Case Study Offtaker Physical Match Ranking",
        "{{DATE}}": REPORT_DATE,
        "{{PROJECT}}": PROJECT_NAME,
        "{{REPO}}": REPO_NAME,
        "{{INPUT_OUTPUT_CONTENT}}": input_output_html,
        "{{MERMAID_DIAGRAM}}": """flowchart TD
A[Six case-study load files] --> B[Parse CSV XLSX and JSON loads]
B --> C{Dirty hours present?}
C -- Yes --> D[Interpolate missing cells and clip negative loads to zero]
C -- No --> E[Keep cleaned 8760 loads as-is]
D --> F[Replay common 30 MWp Ninh Thuan solar profile]
E --> F
F --> G[Compute direct match and 6 MW BESS proxy match]
G --> H[Score physical absorption only]
H --> I[Rank all six candidates and publish JSON plus HTML report]""",
        "{{TOOLS_METHODS}}": tools_html,
        "{{CHARTS_SECTION}}": charts_html,
        "{{OPEN_QUESTIONS}}": "<ul>"
        + "".join(f"<li>{item}</li>" for item in ranking_bullets)
        + "</ul>"
        + open_questions_html,
    }

    html = template_text
    for token, value in replacements.items():
        html = html.replace(token, value)
    return html


def write_outputs(
    results: dict, json_out: Path, html_out: Path, template_path: Path | None
) -> tuple[Path, Path | None]:
    json_out.parent.mkdir(parents=True, exist_ok=True)
    json_out.write_text(
        json.dumps(results, indent=2, ensure_ascii=True), encoding="utf-8"
    )

    html_path = None
    if template_path is not None:
        html_out.parent.mkdir(parents=True, exist_ok=True)
        template_text = template_path.read_text(encoding="utf-8")
        html_out.write_text(build_report_html(results, template_text), encoding="utf-8")
        html_path = html_out

    return json_out, html_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rank Vietnam case studies for pure physical solar-offtaker fit"
    )
    parser.add_argument(
        "--json-out",
        default=str(DEFAULT_JSON_OUT),
        help="Output path for the machine-readable ranking JSON artifact",
    )
    parser.add_argument(
        "--html-out",
        default=str(DEFAULT_HTML_OUT),
        help="Output path for the report-style HTML artifact",
    )
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="Report template path; pass an empty string to skip HTML generation",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    template_path = Path(args.template) if args.template else None
    if template_path is not None and not template_path.exists():
        raise FileNotFoundError(
            f"Report template not found: {template_path}. Pass --template '' to skip HTML generation."
        )

    results = build_results()
    json_path, html_path = write_outputs(
        results=results,
        json_out=Path(args.json_out),
        html_out=Path(args.html_out),
        template_path=template_path,
    )

    print(f"Wrote ranking JSON: {json_path}")
    if html_path is not None:
        print(f"Wrote ranking HTML: {html_path}")


if __name__ == "__main__":
    main()
